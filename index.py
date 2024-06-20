# Python program to scrape website
# and save quotes from website

'''import requests
from bs4 import BeautifulSoup
url = "https://www.amazon.in/dp/B01CSFWXNQ?ref_=cm_sw_r_cp_ud_dp_CDWDPHNXKMMRTSEQCF25&th=1"

payload = {}
headers = {
    'device-memory': '8',
    'sec-ch-device-memory': '8',
    'dpr': '1.3875000476837158',
    'sec-ch-dpr': '1.3875000476837158',
    'viewport-width': '1384',
    'sec-ch-viewport-width': '1384',
    'rtt': '150',
    'downlink': '10',
    'ect': '4g',
    'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-ch-ua-platform-version': '"10.0.0"',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Service-Worker-Navigation-Preload': 'true',
    'Sec-Fetch-Site': 'same-origin',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-User': '?1',
    'Sec-Fetch-Dest': 'document',
    # 'Cookie': 'i18n-prefs=INR; session-id=260-1836480-7894804; session-id-time=2082787201l; ubid-acbin=262-6352960-1631813'
}

response = requests.request("GET", url, headers=headers, data=payload)


# URL = "https://www.amazon.in/dp/B01CSFWXNQ?ref_=cm_sw_r_cp_ud_dp_CDWDPHNXKMMRTSEQCF25"
# URL = "https://www.amazon.in/dp/B01CSFWXNQ?ref_=cm_sw_r_cp_ud_dp_CDWDPHNXKMMRTSEQCF25&th=1"
# r = requests.get(URL)
# print(r.content)

soup = BeautifulSoup(response.content, 'html5lib')
print(soup.prettify())

# products = []  # a list to store quotes

# table = soup.find('div', attrs={'id': 'productTitle'})

# for row in table.findAll('div',
#                          attrs={'class': 'a-size-large product-title-word-break'}):
#     product = {}
#     product['name'] = row.text

#     products.append(product)

# filename = 'dips.csv'
# with open(filename, 'w', newline='') as f:
#     w = csv.DictWriter(f, ['name'])
#     w.writeheader()
#     for product in products:
#         w.writerow(product)


# product_title = soup.find('span', attrs={'id': 'productTitle'})
# print(product_title)
# if product_title:
#     product_name = product_title.get_text(strip=True)
#     products.append({"name": product_name})

# filename = 'dips.csv'
# with open(filename, 'w', newline='') as f:
#     w = csv.DictWriter(f, ['name'])
#     w.writeheader()
#     w.writerows(products)'''
import os
import time
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Create images directory if it doesn't exist
images_dir = './images'
if not os.path.exists(images_dir):
    os.makedirs(images_dir)

# Function to delay execution
def delay(seconds):
    time.sleep(seconds)

# Function to create Amazon search URL from search key
def create_url(search_key):
    return f"https://www.amazon.in/s?k={search_key.replace(' ', '+')}&ref=nb_sb_noss_2"

# Function to scrape product details
def scrape_product(driver):
    product_details = {}

    # Wait for and select the container holding the image thumbnails
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "altImages")))
    thumbnails = driver.find_elements(By.CSS_SELECTOR, ".a-spacing-small.item.imageThumbnail.a-declarative")

    # Extract product name and brand
    product_details['name'] = driver.find_element(By.ID, "productTitle").text.strip()
    product_details['brand'] = driver.find_element(By.ID, "bylineInfo").text.strip()

    # Download and save each product image
    product_details['images'] = []
    for index, thumbnail in enumerate(thumbnails):
        thumbnail.click()  # Click on thumbnail to load the image
        delay(1)  # Wait for the image to load

        # Extract the image URL
        image = driver.find_element(By.CSS_SELECTOR, ".a-dynamic-image.a-stretch-horizontal, .a-dynamic-image.a-stretch-vertical")
        image_url = image.get_attribute('src')

        # Download the image
        response = requests.get(image_url)
        image_path = f"{images_dir}/p_image_{index + 1}.jpg"
        with open(image_path, 'wb') as file:
            file.write(response.content)

        product_details['images'].append(image_path)

    return product_details

# Main script
if __name__ == "__main__":
    search_key = "Peanut Butter"
    url = create_url(search_key)

    # Set up the Selenium WebDriver
    driver = webdriver.Chrome()
    driver.get(url)

    # Select a product and navigate to its page
    # This part might need to be adjusted based on the actual page structure
    product_link = driver.find_element(By.CSS_SELECTOR, 'a-link-normal.s-no-outline')
    product_link.click()

    # Scrape product details
    product = scrape_product(driver)

    # Close the browser
    driver.quit()

    # Create an Excel workbook and add product details and images
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Details"

    ws['A1'] = "Product Brand"
    ws['B1'] = product['brand']
    ws['A2'] = "Product Name"
    ws['B2'] = product['name']

    for index, image_path in enumerate(product['images']):
        img = Image(image_path)
        ws.add_image(img, f'C{index + 1}')

    # Save the workbook
    wb.save("product_details.xlsx")
